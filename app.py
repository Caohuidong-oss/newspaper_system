from flask import Flask, render_template, request, redirect, url_for, flash, session
from config import Config
from models import db, User, Newspaper, Order, Subscription, LoginUser
from sqlalchemy import func
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from flask_wtf.csrf import CSRFProtect
import os
from datetime import datetime, timedelta
from api import api

app = Flask(__name__)
app.config.from_object(Config)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # 禁用静态文件缓存
db.init_app(app)
csrf = CSRFProtect(app)
app.register_blueprint(api)
csrf.exempt(api)  # API 蓝图豁免 CSRF（小程序/外部调用）

# 让 HTML 响应不缓存，避免浏览器拿到老模板
@app.after_request
def add_no_cache_headers(response):
    if response.content_type and 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ==================== 图片上传配置 ====================
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_newspaper_image(file):
    """保存上传的报刊图片，返回保存后的文件名"""
    if file and file.filename and allowed_file(file.filename):
        # 生成唯一文件名避免冲突
        import uuid
        ext = file.filename.rsplit('.', 1)[1].lower()
        unique_name = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, unique_name))
        return unique_name
    return None

# ==================== 权限装饰器 ====================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            flash('请先登录', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            flash('请先登录', 'warning')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('需要管理员权限', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== 辅助函数 ====================
def get_current_user():
    """获取当前登录用户对应的订户记录"""
    username = session.get('username')
    if not username:
        return None
    return User.query.filter_by(username=username).first()

# ==================== 认证路由 ====================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = LoginUser.query.filter_by(username=username).first()
        if not user:
            flash('该用户不存在', 'danger')
        elif check_password_hash(user.password_hash, password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            flash(f'欢迎回来，{user.username}！', 'success')
            next_url = request.args.get('next') or url_for('index')
            return redirect(next_url)
        else:
            flash('密码错误', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm = request.form.get('confirm_password')
        # 安全：忽略前端传值，强制写死为 user（防止越权注册为管理员）
        role = 'user'
        
        if not username or not password:
            flash('用户名和密码不能为空', 'danger')
            return render_template('register.html')
        if password != confirm:
            flash('两次密码输入不一致', 'danger')
            return render_template('register.html')
        
        existing = LoginUser.query.filter_by(username=username).first()
        if existing:
            flash('用户名已存在，请更换', 'danger')
            return render_template('register.html')
        
        # 创建登录账号
        new_user = LoginUser(
            username=username,
            password_hash=generate_password_hash(password),
            role=role
        )
        db.session.add(new_user)
        db.session.flush()
        
        # 自动为订阅者创建对应的订户记录
        existing_user = User.query.filter_by(username=username).first()
        if not existing_user:
            new_subscriber = User(
                username=username,
                password='',  # 密码已存于 LoginUser.password_hash，此处不再存明文
                real_name=username,
                phone='',
                address=''
            )
            db.session.add(new_subscriber)
        
        db.session.commit()
        flash('注册成功！现在您可以直接订阅报刊了', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('已退出登录', 'info')
    return redirect(url_for('index'))

@app.route('/api_test')
def api_test():
    """API 测试页（开发用，方便调试 JWT）"""
    return render_template('api_test.html')
    return redirect(url_for('index'))

# ==================== 个人中心 ====================
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    subscriber = get_current_user()

    # 如果没有订户记录，自动创建一条
    if not subscriber:
        username = session.get('username')
        subscriber = User(
            username=username,
            password='N/A',  # 占位密码：admin 用户用 LoginUser 登录，此字段不会被使用
            real_name=username,
            phone='',
            address=''
        )
        db.session.add(subscriber)
        db.session.commit()

    if request.method == 'POST':
        subscriber.real_name = request.form.get('real_name', '').strip()
        subscriber.phone = request.form.get('phone', '').strip()
        subscriber.address = request.form.get('address', '').strip()

        if not subscriber.real_name:
            flash('真实姓名不能为空', 'danger')
            return render_template('profile.html', subscriber=subscriber)

        db.session.commit()
        flash('个人信息更新成功！', 'success')
        return redirect(url_for('profile'))

    # 简单统计：累计订单数和总消费
    order_count = Order.query.filter_by(user_id=subscriber.user_id).count()
    total_spent = db.session.query(func.sum(Order.total_amount))\
        .filter(Order.user_id == subscriber.user_id).scalar() or 0

    # 管理员全局概览
    admin_stats = {}
    if session.get('role') == 'admin':
        yesterday = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
        yesterday_end = yesterday + timedelta(hours=23, minutes=59, seconds=59)
        today_start = yesterday + timedelta(days=1)

        admin_stats = {
            'total_users': User.query.count(),
            'total_newspapers': Newspaper.query.count(),
            'total_orders': Order.query.count(),
            'total_revenue': float(db.session.query(func.sum(Order.total_amount)).scalar() or 0),
            'yesterday_new_users': User.query.filter(User.register_date.between(yesterday, yesterday_end)).count(),
            'yesterday_new_orders': Order.query.filter(Order.order_date.between(yesterday, yesterday_end)).count(),
            'today_new_users': User.query.filter(User.register_date >= today_start).count(),
            'today_new_orders': Order.query.filter(Order.order_date >= today_start).count(),
        }

    return render_template('profile.html',
                           subscriber=subscriber,
                           order_count=order_count,
                           total_spent=total_spent,
                           admin_stats=admin_stats)

# ==================== 修改密码 ====================
@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        login_user = LoginUser.query.get(session.get('user_id'))
        if not login_user or not check_password_hash(login_user.password_hash, old_password):
            flash('原密码错误', 'danger')
            return render_template('change_password.html')

        if not new_password or len(new_password) < 6:
            flash('新密码长度不能少于6位', 'danger')
            return render_template('change_password.html')

        if new_password != confirm_password:
            flash('两次输入的新密码不一致', 'danger')
            return render_template('change_password.html')

        login_user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('密码修改成功！请使用新密码重新登录', 'success')
        session.clear()
        return redirect(url_for('login'))

    return render_template('change_password.html')

# ==================== 找回密码（安全问题方案） ====================
SECURITY_QUESTIONS = [
    '您的出生城市是？',
    '您母亲的姓名是？',
    '您小学班主任的姓名是？',
    '您最喜欢的食物是？',
    '您的第一只宠物叫什么名字？',
    '您的童年昵称是？',
]

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """三步式找回：step1=输入用户名, step2=答安全问题, step3=重置密码"""
    step = request.form.get('step', '1')
    username = request.form.get('username', '').strip()

    # Step 1: 输入用户名
    if step == '1':
        if request.method == 'POST':
            if not username:
                flash('请输入用户名', 'danger')
                return render_template('forgot_password.html', step='1')
            user = LoginUser.query.filter_by(username=username).first()
            if not user:
                flash('该用户不存在', 'danger')
                return render_template('forgot_password.html', step='1')
            if not user.security_question:
                flash('该账户未设置安全问题，请联系管理员重置密码', 'warning')
                return render_template('forgot_password.html', step='1')
            return render_template('forgot_password.html', step='2', username=username, question=user.security_question)
        return render_template('forgot_password.html', step='1')

    # Step 2: 回答安全问题
    if step == '2':
        if request.method == 'POST':
            answer = request.form.get('answer', '').strip()
            user = LoginUser.query.filter_by(username=username).first()
            if not user or not check_password_hash(user.security_answer_hash, answer.lower()):
                flash('答案错误', 'danger')
                return render_template('forgot_password.html', step='2', username=username, question=user.security_question if user else '')
            return render_template('forgot_password.html', step='3', username=username)
        return render_template('forgot_password.html', step='2', username=username)

    # Step 3: 设置新密码
    if step == '3':
        if request.method == 'POST':
            new_password = request.form.get('new_password', '')
            confirm_password = request.form.get('confirm_password', '')
            if not new_password or len(new_password) < 6:
                flash('新密码长度不能少于6位', 'danger')
                return render_template('forgot_password.html', step='3', username=username)
            if new_password != confirm_password:
                flash('两次密码输入不一致', 'danger')
                return render_template('forgot_password.html', step='3', username=username)
            user = LoginUser.query.filter_by(username=username).first()
            if not user:
                flash('用户不存在', 'danger')
                return render_template('forgot_password.html', step='1')
            user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            flash('密码重置成功！请使用新密码登录', 'success')
            return redirect(url_for('login'))
        return render_template('forgot_password.html', step='3', username=username)

    return render_template('forgot_password.html', step='1')

# ==================== 设置安全问题 ====================
@app.route('/set_security', methods=['GET', 'POST'])
@login_required
def set_security():
    """设置/更新当前用户的安全问题"""
    login_user = LoginUser.query.get(session.get('user_id'))

    if request.method == 'POST':
        question = request.form.get('question', '').strip()
        answer = request.form.get('answer', '').strip().lower()

        if not question or question not in SECURITY_QUESTIONS:
            flash('请选择有效的安全问题', 'danger')
            return render_template('set_security.html', login_user=login_user, questions=SECURITY_QUESTIONS)
        if not answer or len(answer) < 2:
            flash('答案至少2个字符', 'danger')
            return render_template('set_security.html', login_user=login_user, questions=SECURITY_QUESTIONS)

        login_user.security_question = question
        login_user.security_answer_hash = generate_password_hash(answer)
        db.session.commit()
        flash('安全问题设置成功！', 'success')
        return redirect(url_for('set_security'))

    return render_template('set_security.html', login_user=login_user, questions=SECURITY_QUESTIONS)

# ==================== 报刊详情 ====================
@app.route('/newspaper/detail/<int:newspaper_id>')
@login_required
def newspaper_detail(newspaper_id):
    newspaper = Newspaper.query.get_or_404(newspaper_id)

    # 订阅统计
    sub_stats = db.session.query(
        func.sum(Subscription.quantity).label('total_qty'),
        func.sum(Subscription.subtotal).label('total_amount')
    ).filter(Subscription.newspaper_id == newspaper_id).first()

    total_qty = int(sub_stats.total_qty or 0)
    total_amount = float(sub_stats.total_amount or 0)

    # 该报刊的所有订单
    orders = db.session.query(Order, Subscription, User)\
        .join(Subscription, Subscription.order_id == Order.order_id)\
        .join(User, User.user_id == Order.user_id)\
        .filter(Subscription.newspaper_id == newspaper_id)\
        .order_by(Order.order_date.desc()).limit(20).all()

    return render_template('newspaper_detail.html',
                           newspaper=newspaper,
                           total_qty=total_qty,
                           total_amount=total_amount,
                           orders=orders)

# ==================== 首页 ====================
@app.route('/')
def index():
    is_admin = session.get('role') == 'admin'
    username = session.get('username')

    if is_admin:
        # 管理员：看全平台数据
        total_users = User.query.count()
        total_orders = Order.query.count()
        total_revenue = db.session.query(func.sum(Subscription.subtotal)).scalar() or 0
        recent_orders = Order.query.order_by(Order.order_id.desc()).limit(5).all()
    else:
        # 订阅者：只看自己的数据
        current_user = User.query.filter_by(username=username).first() if username else None
        if current_user:
            total_users = 1  # 只有自己
            total_orders = Order.query.filter_by(user_id=current_user.user_id).count()
            total_revenue = db.session.query(func.sum(Order.total_amount))\
                .filter(Order.user_id == current_user.user_id).scalar() or 0
            recent_orders = Order.query.filter_by(user_id=current_user.user_id)\
                             .order_by(Order.order_id.desc()).limit(5).all()
        else:
            total_users = 0
            total_orders = 0
            total_revenue = 0
            recent_orders = []

    total_newspapers = Newspaper.query.count()

    top_newspapers = db.session.query(
        Newspaper.name,
        func.sum(Subscription.quantity).label('total_qty')
    ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
     .group_by(Newspaper.newspaper_id)\
     .order_by(func.sum(Subscription.quantity).desc())\
     .limit(5).all()

    return render_template('index.html',
        is_admin=is_admin,
        total_users=total_users,
        total_newspapers=total_newspapers,
        total_orders=total_orders,
        total_revenue=total_revenue,
        recent_orders=recent_orders,
        top_newspapers=top_newspapers
    )

# ==================== 订户管理（仅管理员） ====================
@app.route('/users')
@login_required
def users():
    if session.get('role') != 'admin':
        flash('您无权访问此页面', 'danger')
        return redirect(url_for('index'))

    keyword = request.args.get('keyword', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20

    if keyword:
        query = User.query.filter(
            User.username.contains(keyword) | User.real_name.contains(keyword)
        )
    else:
        query = User.query
    pagination = query.order_by(User.user_id).paginate(page=page, per_page=per_page, error_out=False)
    all_users = pagination.items

    return render_template('users.html', users=all_users,
                           pagination=pagination, keyword=keyword)

# 订户必须自己注册，不允许管理员代为创建
@app.route('/user/add', methods=['GET', 'POST'])
@admin_required
def user_add():
    flash('订户需自行注册账号，请引导用户访问注册页面', 'info')
    return redirect(url_for('register'))

@app.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def user_edit(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.username = request.form.get('username')
        user.real_name = request.form.get('real_name')
        user.phone = request.form.get('phone', '')
        user.address = request.form.get('address', '')
        new_password = request.form.get('password')
        if new_password:
            # 同步更新登录密码（写入 LoginUser 的哈希字段）
            login_user = LoginUser.query.filter_by(username=user.username).first()
            if login_user:
                login_user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        flash('订户信息更新成功！', 'success')
        return redirect(url_for('users'))
    return render_template('user_form.html', user=user)

@app.route('/user/delete/<int:user_id>', methods=['POST'])
@admin_required
def user_delete(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('订户已删除！', 'danger')
    return redirect(url_for('users'))

@app.route('/user/detail/<int:user_id>')
@admin_required
def user_detail(user_id):
    """订户详情：基本信息 + 所有订单 + 订阅统计"""
    user = User.query.get_or_404(user_id)
    orders = Order.query.filter_by(user_id=user_id).order_by(Order.order_id.desc()).all()
    order_count = len(orders)
    total_spent = db.session.query(func.sum(Order.total_amount))\
        .filter(Order.user_id == user_id).scalar() or 0

    # 订阅偏好：该用户订阅最多的报刊 Top 3
    top_newspapers = db.session.query(
        Newspaper.name,
        func.sum(Subscription.quantity).label('total_qty')
    ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
     .join(Order, Order.order_id == Subscription.order_id)\
     .filter(Order.user_id == user_id)\
     .group_by(Newspaper.newspaper_id)\
     .order_by(func.sum(Subscription.quantity).desc())\
     .limit(3).all()

    return render_template('user_detail.html',
                           user=user,
                           orders=orders,
                           order_count=order_count,
                           total_spent=total_spent,
                           top_newspapers=top_newspapers)

# ==================== 管理员管理 ====================
@app.route('/admin/users')
@admin_required
def admin_users():
    """管理员管理：查看所有登录用户及其角色"""
    login_users = LoginUser.query.order_by(LoginUser.id).all()
    return render_template('admin_users.html', login_users=login_users)

@app.route('/admin/set_role/<int:user_id>', methods=['POST'])
@admin_required
def admin_set_role(user_id):
    """切换用户的角色（user ↔ admin）"""
    user = LoginUser.query.get_or_404(user_id)
    if user.id == session.get('user_id'):
        flash('不能取消自己的管理员身份', 'danger')
        return redirect(url_for('admin_users'))

    user.role = 'admin' if user.role == 'user' else 'user'
    role_name = '管理员' if user.role == 'admin' else '订阅者'
    db.session.commit()
    flash(f'用户「{user.username}」已设为{role_name}', 'success')
    return redirect(url_for('admin_users'))

# ==================== 报刊管理 ====================
@app.route('/newspapers')
@login_required
def newspapers():
    keyword = request.args.get('keyword', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 12  # 报刊每页少一点，卡片布局更舒服

    if keyword:
        query = Newspaper.query.filter(Newspaper.name.contains(keyword))
    else:
        query = Newspaper.query
    pagination = query.order_by(Newspaper.newspaper_id).paginate(page=page, per_page=per_page, error_out=False)
    all_newspapers = pagination.items

    return render_template('newspapers.html', newspapers=all_newspapers,
                           pagination=pagination, keyword=keyword)

@app.route('/newspaper/add', methods=['GET', 'POST'])
@admin_required
def newspaper_add():
    if request.method == 'POST':
        name = request.form.get('name')
        type_ = request.form.get('type', '')
        price = request.form.get('price')
        period = request.form.get('period', '')
        description = request.form.get('description', '')
        
        # 处理图片上传
        image = save_newspaper_image(request.files.get('image'))
        
        new_newspaper = Newspaper(
            name=name,
            type=type_,
            price=price,
            period=period,
            description=description,
            image=image or ''
        )
        db.session.add(new_newspaper)
        db.session.commit()
        flash('报刊添加成功！', 'success')
        return redirect(url_for('newspapers'))
    return render_template('newspaper_form.html', newspaper=None)

@app.route('/newspaper/edit/<int:newspaper_id>', methods=['GET', 'POST'])
@admin_required
def newspaper_edit(newspaper_id):
    newspaper = Newspaper.query.get_or_404(newspaper_id)
    if request.method == 'POST':
        newspaper.name = request.form.get('name')
        newspaper.type = request.form.get('type', '')
        newspaper.price = request.form.get('price')
        newspaper.period = request.form.get('period', '')
        newspaper.description = request.form.get('description', '')
        
        # 处理图片上传（有新图则替换）
        uploaded = save_newspaper_image(request.files.get('image'))
        if uploaded:
            # 删除旧图片
            if newspaper.image:
                old_path = os.path.join(UPLOAD_FOLDER, newspaper.image)
                if os.path.exists(old_path):
                    os.remove(old_path)
            newspaper.image = uploaded
        
        db.session.commit()
        flash('报刊信息更新成功！', 'success')
        return redirect(url_for('newspapers'))
    return render_template('newspaper_form.html', newspaper=newspaper)

@app.route('/newspaper/delete/<int:newspaper_id>', methods=['POST'])
@admin_required
def newspaper_delete(newspaper_id):
    newspaper = Newspaper.query.get_or_404(newspaper_id)
    # 删除关联的图片文件
    if newspaper.image:
        img_path = os.path.join(UPLOAD_FOLDER, newspaper.image)
        if os.path.exists(img_path):
            os.remove(img_path)
    db.session.delete(newspaper)
    db.session.commit()
    flash('报刊已删除！', 'danger')
    return redirect(url_for('newspapers'))

# ==================== 订单管理 ====================
@app.route('/orders')
@login_required
def orders():
    keyword = request.args.get('keyword', '').strip()
    status_filter = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = 20

    if session.get('role') == 'admin':
        query = Order.query
        if keyword:
            query = query.join(User).filter(User.real_name.contains(keyword))
        if status_filter and status_filter.isdigit():
            query = query.filter(Order.status == int(status_filter))
        pagination = query.order_by(Order.order_id.desc()).paginate(page=page, per_page=per_page, error_out=False)
        all_orders = pagination.items
    else:
        user = User.query.filter_by(username=session.get('username')).first()
        if user:
            query = Order.query.filter_by(user_id=user.user_id)
            if status_filter and status_filter.isdigit():
                query = query.filter(Order.status == int(status_filter))
            pagination = query.order_by(Order.order_id.desc()).paginate(page=page, per_page=per_page, error_out=False)
            all_orders = pagination.items
        else:
            all_orders = []
            pagination = None

    return render_template('orders.html', orders=all_orders,
                           pagination=pagination,
                           keyword=keyword, status_filter=status_filter)

@app.route('/order/detail/<int:order_id>')
@login_required
def order_detail(order_id):
    order = Order.query.get_or_404(order_id)

    if session.get('role') != 'admin':
        user = User.query.filter_by(username=session.get('username')).first()
        if not user or order.user_id != user.user_id:
            flash('您无权查看此订单', 'danger')
            return redirect(url_for('orders'))

    return render_template('order_detail.html', order=order)

@app.route('/order/print/<int:order_id>')
@login_required
def order_print(order_id):
    """订单打印页：纯净版，去除导航栏和按钮"""
    order = Order.query.get_or_404(order_id)

    if session.get('role') != 'admin':
        user = User.query.filter_by(username=session.get('username')).first()
        if not user or order.user_id != user.user_id:
            flash('您无权查看此订单', 'danger')
            return redirect(url_for('orders'))

    return render_template('order_print.html', order=order)

@app.route('/order/add', methods=['GET', 'POST'])
@login_required
def order_add():
    # 管理员不需要订阅报刊
    if session.get('role') == 'admin':
        flash('管理员无需订阅报刊', 'info')
        return redirect(url_for('orders'))

    current_user = User.query.filter_by(username=session.get('username')).first()

    if not current_user:
        flash('您还没有订户账号，请联系管理员添加', 'warning')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if session.get('role') != 'admin':
            user_id = current_user.user_id
        else:
            user_id = request.form.get('user_id')
            if not user_id:
                flash('请选择订户', 'danger')
                return redirect(url_for('order_add'))
        
        items = []
        total = 0
        for key, value in request.form.items():
            if key.startswith('qty_'):
                newspaper_id = int(key.split('_')[1])
                qty = int(value) if value else 0
                if qty > 0:
                    newspaper = Newspaper.query.get(newspaper_id)
                    subtotal = newspaper.price * qty
                    items.append({
                        'newspaper_id': newspaper_id,
                        'quantity': qty,
                        'subtotal': subtotal
                    })
                    total += subtotal
        if not items:
            flash('请至少选择一种报刊并填写数量', 'warning')
            return redirect(url_for('order_add'))
        
        order = Order(user_id=user_id, total_amount=total, status=1,
                      note=request.form.get('note', '').strip())
        db.session.add(order)
        db.session.flush()
        for item in items:
            sub = Subscription(
                order_id=order.order_id,
                newspaper_id=item['newspaper_id'],
                quantity=item['quantity'],
                subtotal=item['subtotal']
            )
            db.session.add(sub)
        db.session.commit()
        flash('订单创建成功！', 'success')
        return redirect(url_for('orders'))
    
    users = User.query.all() if session.get('role') == 'admin' else [current_user] if current_user else []
    newspapers = Newspaper.query.all()
    return render_template('order_form.html', users=users, newspapers=newspapers)

@app.route('/order/cancel/<int:order_id>', methods=['POST'])
@login_required
def order_cancel(order_id):
    order = Order.query.get_or_404(order_id)
    
    if session.get('role') != 'admin':
        user = User.query.filter_by(username=session.get('username')).first()
        if not user or order.user_id != user.user_id:
            flash('您无权操作此订单', 'danger')
            return redirect(url_for('orders'))
        if order.status != 1:
            flash('只有待处理的订单可以取消', 'warning')
            return redirect(url_for('order_detail', order_id=order_id))
    
    order.status = 3
    db.session.commit()
    flash('订单已取消', 'warning')
    return redirect(url_for('orders'))

@app.route('/order/delete/<int:order_id>', methods=['POST'])
@admin_required
def order_delete(order_id):
    order = Order.query.get_or_404(order_id)
    db.session.delete(order)
    db.session.commit()
    flash('订单已删除', 'danger')
    return redirect(url_for('orders'))

@app.route('/order/clean_empty', methods=['POST'])
@admin_required
def order_clean_empty():
    """清理空订单（无订阅明细的订单）
    已弃用：前端已强制要求至少选一种报刊才允许下单，不会产生空订单。
    保留仅为兼容性。
    """
    return redirect(url_for('orders'))

@app.route('/order/batch_action', methods=['POST'])
@admin_required
def order_batch_action():
    """批量操作订单：批量确认/取消/删除"""
    action = request.form.get('action')
    order_ids = request.form.getlist('order_ids')

    if not order_ids:
        flash('请先勾选要操作的订单', 'warning')
        return redirect(url_for('orders'))

    # 转 int 过滤无效 ID
    try:
        order_ids = [int(i) for i in order_ids]
    except (ValueError, TypeError):
        flash('参数错误', 'danger')
        return redirect(url_for('orders'))

    orders = Order.query.filter(Order.order_id.in_(order_ids)).all()
    count = len(orders)

    if action == 'confirm':
        n = 0
        for o in orders:
            if o.status == 1:
                o.status = 2
                n += 1
        db.session.commit()
        flash(f'已批量确认 {n} 条订单', 'success')
    elif action == 'cancel':
        n = 0
        for o in orders:
            if o.status != 3:
                o.status = 3
                n += 1
        db.session.commit()
        flash(f'已批量取消 {n} 条订单', 'warning')
    elif action == 'delete':
        for o in orders:
            db.session.delete(o)
        db.session.commit()
        flash(f'已批量删除 {count} 条订单', 'danger')
    else:
        flash('未知操作', 'danger')

    return redirect(url_for('orders'))

@app.route('/order/confirm/<int:order_id>', methods=['POST'])
@admin_required
def order_confirm(order_id):
    """管理员确认订单：待处理 → 已确认"""
    order = Order.query.get_or_404(order_id)
    if order.status != 1:
        flash('只有待处理的订单可以确认', 'warning')
        return redirect(url_for('order_detail', order_id=order_id))
    order.status = 2
    db.session.commit()
    flash('订单已确认 ✅', 'success')
    return redirect(url_for('orders'))

# ==================== 订单查询 ====================
@app.route('/order/query', methods=['GET', 'POST'])
@login_required
def order_query():
    orders = []
    keyword = ''
    
    if request.method == 'POST':
        keyword = request.form.get('keyword', '').strip()
        
        if session.get('role') == 'admin':
            if keyword:
                orders = Order.query.join(User).filter(
                    User.real_name.like(f'%{keyword}%')
                ).order_by(Order.order_id.desc()).all()
            else:
                orders = Order.query.order_by(Order.order_id.desc()).all()
        else:
            user = User.query.filter_by(username=session.get('username')).first()
            if user:
                orders = Order.query.filter_by(user_id=user.user_id).order_by(Order.order_id.desc()).all()
    else:
        if session.get('role') == 'admin':
            orders = Order.query.order_by(Order.order_id.desc()).all()
        else:
            user = User.query.filter_by(username=session.get('username')).first()
            if user:
                orders = Order.query.filter_by(user_id=user.user_id).order_by(Order.order_id.desc()).all()
    
    return render_template('order_query.html', orders=orders, keyword=keyword)

# ==================== 导出 Excel ====================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime as dt

def make_excel_response(filename, headers, rows):
    """生成 Excel 文件并返回响应"""
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'

    # 标题行样式
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    # 写标题
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 写数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 自动列宽
    for col_idx, h in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(h))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # 保存到内存
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export/orders')
@login_required
def export_orders():
    """导出订单列表为 Excel"""
    if session.get('role') == 'admin':
        orders = Order.query.order_by(Order.order_id.desc()).all()
    else:
        user = get_current_user()
        if user:
            orders = Order.query.filter_by(user_id=user.user_id).order_by(Order.order_id.desc()).all()
        else:
            orders = []

    headers = ['订单号', '订户', '用户名', '下单时间', '商品数', '总金额(元)', '状态']
    status_map = {1: '待处理', 2: '已确认', 3: '已取消'}
    rows = []
    for o in orders:
        total_qty = sum(sub.quantity for sub in o.subscriptions)
        rows.append([
            o.order_id,
            o.user.real_name,
            o.user.username,
            o.order_date.strftime('%Y-%m-%d %H:%M:%S'),
            total_qty,
            float(o.total_amount),
            status_map.get(o.status, '未知')
        ])

    filename = f"订单列表_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(filename, headers, rows)

@app.route('/export/users')
@admin_required
def export_users():
    """导出订户列表为 Excel"""
    users = User.query.order_by(User.user_id).all()
    headers = ['订户ID', '用户名', '真实姓名', '电话', '地址', '注册日期']
    rows = [[
        u.user_id, u.username, u.real_name,
        u.phone or '', u.address or '',
        u.register_date.strftime('%Y-%m-%d') if u.register_date else ''
    ] for u in users]

    filename = f"订户列表_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(filename, headers, rows)

@app.route('/export/newspapers')
@login_required
def export_newspapers():
    """导出报刊列表为 Excel"""
    newspapers = Newspaper.query.order_by(Newspaper.newspaper_id).all()
    headers = ['ID', '报刊名称', '类型', '单价(元)', '出版周期', '简介']
    rows = [[
        n.newspaper_id, n.name, n.type or '',
        float(n.price), n.period or '', n.description or ''
    ] for n in newspapers]

    filename = f"报刊列表_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_excel_response(filename, headers, rows)

@app.route('/export/statistics')
@login_required
def export_statistics():
    """导出统计报表为 Excel（多 sheet）"""
    wb = Workbook()

    # ===== Sheet 1: 报刊统计 =====
    is_admin = session.get('role') == 'admin'
    if is_admin:
        stats = db.session.query(
            Newspaper.name,
            func.sum(Subscription.quantity).label('total_quantity'),
            func.sum(Subscription.subtotal).label('total_amount')
        ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
         .group_by(Newspaper.newspaper_id)\
         .order_by(func.sum(Subscription.subtotal).desc()).all()
    else:
        user = get_current_user()
        if user:
            stats = db.session.query(
                Newspaper.name,
                func.sum(Subscription.quantity).label('total_quantity'),
                func.sum(Subscription.subtotal).label('total_amount')
            ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
             .filter(Subscription.order.has(user_id=user.user_id))\
             .group_by(Newspaper.newspaper_id)\
             .order_by(func.sum(Subscription.subtotal).desc()).all()
        else:
            stats = []

    ws1 = wb.active
    ws1.title = '报刊统计'
    ws1.append(['报刊名称', '订阅份数', '销售金额(元)'])
    for s in stats:
        ws1.append([s.name, s.total_quantity, float(s.total_amount)])

    # 设置样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws1.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ===== Sheet 2: 类型统计 =====
    if is_admin:
        type_stats = db.session.query(
            Newspaper.type,
            func.sum(Subscription.quantity).label('total_quantity'),
            func.sum(Subscription.subtotal).label('total_amount')
        ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
         .group_by(Newspaper.type)\
         .order_by(func.sum(Subscription.subtotal).desc()).all()
    else:
        user = get_current_user()
        type_stats = []
        if user:
            type_stats = db.session.query(
                Newspaper.type,
                func.sum(Subscription.quantity).label('total_quantity'),
                func.sum(Subscription.subtotal).label('total_amount')
            ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
             .filter(Subscription.order.has(user_id=user.user_id))\
             .group_by(Newspaper.type)\
             .order_by(func.sum(Subscription.subtotal).desc()).all()

    ws2 = wb.create_sheet('类型统计')
    ws2.append(['类型', '总份数', '总金额(元)'])
    for t in type_stats:
        ws2.append([t.type or '未分类', t.total_quantity, float(t.total_amount)])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws2.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    filename = f"统计报表_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==================== 统计查询 ====================
@app.route('/statistics')
@login_required
def statistics():
    if session.get('role') == 'admin':
        stats = db.session.query(
            Newspaper.name,
            func.sum(Subscription.quantity).label('total_quantity'),
            func.sum(Subscription.subtotal).label('total_amount')
        ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
         .group_by(Newspaper.newspaper_id)\
         .order_by(func.sum(Subscription.subtotal).desc()).all()
        total_subscriptions = db.session.query(func.sum(Subscription.quantity)).scalar() or 0
        total_revenue = db.session.query(func.sum(Subscription.subtotal)).scalar() or 0
        total_orders = db.session.query(Order).count()
        total_users = db.session.query(User).count()
        total_newspapers = db.session.query(Newspaper).count()
        type_stats = db.session.query(
            Newspaper.type,
            func.sum(Subscription.quantity).label('total_quantity'),
            func.sum(Subscription.subtotal).label('total_amount')
        ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
         .group_by(Newspaper.type)\
         .order_by(func.sum(Subscription.subtotal).desc()).all()
        chart_labels = [item.name for item in stats]
        chart_data = [float(item.total_amount) for item in stats]
        chart_colors = ['#0d6efd', '#6610f2', '#6f42c1', '#d63384', '#dc3545',
                        '#fd7e14', '#ffc107', '#198754', '#0dcaf0', '#0d6efd']
        max_amount = stats[0].total_amount if stats else 0
    else:
        user = User.query.filter_by(username=session.get('username')).first()
        if user:
            stats = db.session.query(
                Newspaper.name,
                func.sum(Subscription.quantity).label('total_quantity'),
                func.sum(Subscription.subtotal).label('total_amount')
            ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
             .filter(Subscription.order.has(user_id=user.user_id))\
             .group_by(Newspaper.newspaper_id)\
             .order_by(func.sum(Subscription.subtotal).desc()).all()
            total_subscriptions = db.session.query(func.sum(Subscription.quantity)).filter(
                Subscription.order.has(user_id=user.user_id)
            ).scalar() or 0
            total_revenue = db.session.query(func.sum(Subscription.subtotal)).filter(
                Subscription.order.has(user_id=user.user_id)
            ).scalar() or 0
            total_orders = Order.query.filter_by(user_id=user.user_id).count()
            total_users = 1
            total_newspapers = Newspaper.query.count()
            type_stats = db.session.query(
                Newspaper.type,
                func.sum(Subscription.quantity).label('total_quantity'),
                func.sum(Subscription.subtotal).label('total_amount')
            ).join(Subscription, Newspaper.newspaper_id == Subscription.newspaper_id)\
             .filter(Subscription.order.has(user_id=user.user_id))\
             .group_by(Newspaper.type)\
             .order_by(func.sum(Subscription.subtotal).desc()).all()
            chart_labels = [item.name for item in stats]
            chart_data = [float(item.total_amount) for item in stats]
            chart_colors = ['#0d6efd', '#6610f2', '#6f42c1', '#d63384', '#dc3545']
            max_amount = stats[0].total_amount if stats else 0
        else:
            stats = []
            total_subscriptions = 0
            total_revenue = 0
            total_orders = 0
            total_users = 1
            total_newspapers = Newspaper.query.count()
            type_stats = []
            chart_labels = []
            chart_data = []
            chart_colors = []
            max_amount = 0
    
    return render_template('statistics.html',
        stats=stats,
        max_amount=max_amount,
        total_subscriptions=total_subscriptions,
        total_revenue=total_revenue,
        total_orders=total_orders,
        total_users=total_users,
        total_newspapers=total_newspapers,
        type_stats=type_stats,
        chart_labels=chart_labels,
        chart_data=chart_data,
        chart_colors=chart_colors
    )

# ==================== 404 错误处理 ====================
@app.errorhandler(404)
def page_not_found(e):
    return '''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <title>404 - 页面未找到</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0-alpha1/dist/css/bootstrap.min.css" rel="stylesheet">
        <style>
            body { background: linear-gradient(135deg, #667eea, #764ba2); min-height: 100vh; display: flex; align-items: center; }
            .error-card { background: #fff; border-radius: 16px; box-shadow: 0 8px 32px rgba(0,0,0,.2); }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="row justify-content-center">
                <div class="col-md-6">
                    <div class="error-card p-5 text-center">
                        <div class="display-1 text-primary mb-3">🔍</div>
                        <h1 class="display-4 fw-bold text-primary">404</h1>
                        <p class="text-muted mb-4">您访问的页面不存在，可能已被移除</p>
                        <a href="/" class="btn btn-primary rounded-pill px-4">🏠 返回首页</a>
                    </div>
                </div>
            </div>
        </div>
    </body>
    </html>
    ''', 404

# ==================== 启动 ====================
if __name__ == '__main__':
    # Railway 通过 PORT 环境变量指定端口
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
